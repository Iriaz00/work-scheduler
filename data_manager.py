
import json
from typing import Dict, List, Tuple
from models import Employee, FixedSchedule, ShiftType, MonthSchedule, ALL_SHIFTS


def save_config(
    year: int, month: int,
    employees: List[Employee],
    carryover: Dict[str, Dict[int, ShiftType]],
    filepath: str = "config.json",
) -> str:
    data = {
        "year": year, "month": month,
        "employees": [
            {
                "name": e.name,
                "prefer_day": e.prefer_day,
                "prefer_night": e.prefer_night,
                "fixed_schedules": [
                    {"day": fs.day, "shift": fs.shift.value}
                    for fs in e.fixed_schedules
                ],
            }
            for e in employees
        ],
        "carryover": {
            name: {str(d): s.value for d, s in days.items()}
            for name, days in carryover.items()
        },
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return filepath


def load_config(filepath: str = "config.json"):
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    employees = [
        Employee(
            name=e["name"],
            prefer_day=e["prefer_day"],
            prefer_night=e["prefer_night"],
            fixed_schedules=[
                FixedSchedule(day=fs["day"], shift=ShiftType(fs["shift"]))
                for fs in e["fixed_schedules"]
            ],
        )
        for e in data["employees"]
    ]
    carryover = {
        name: {int(d): ShiftType(s) for d, s in days.items()}
        for name, days in data.get("carryover", {}).items()
    }
    return data["year"], data["month"], employees, carryover


def export_excel(result: MonthSchedule, filepath: str = "schedule.xlsx") -> str:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    COLORS = {
        ShiftType.DAY:       ("4A90D9", "FFFFFF"),
        ShiftType.NIGHT:     ("1A237E", "FFFFFF"),
        ShiftType.OFF:       ("90A4AE", "FFFFFF"),
        ShiftType.HOLIDAY:   ("43A047", "FFFFFF"),
        ShiftType.ANNUAL:    ("66BB6A", "FFFFFF"),
        ShiftType.SPECIAL:   ("FFA726", "FFFFFF"),
        ShiftType.EDUCATION: ("AB47BC", "FFFFFF"),
    }

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"{result.year}-{result.month:02d}"
    days = list(range(1, result.num_days + 1))

    # 헤더
    ws.cell(1, 1, "이름").font = Font(bold=True)
    for col, d in enumerate(days, 2):
        c = ws.cell(1, col, str(d))
        c.font, c.alignment = Font(bold=True), Alignment(horizontal="center")
    ws.cell(1, len(days) + 2, "통계").font = Font(bold=True)

    # 직원 행
    for row, emp in enumerate(result.employees, 2):
        ws.cell(row, 1, emp.name).font = Font(bold=True)
        for col, d in enumerate(days, 2):
            shift = result.get_shift(emp.name, d)
            cell  = ws.cell(row, col, shift.value if shift else "")
            cell.alignment = Alignment(horizontal="center")
            if shift and shift in COLORS:
                bg, fg = COLORS[shift]
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                cell.font = Font(color=fg)
        # 통계 열
        st = result.get_stats(emp.name)
        ws.cell(row, len(days) + 2,
                f"주{st['주간']} 야{st['야간']} 비{st['비번']} 휴{st['휴일']} 연{st['연가']}")

    # 합계 행
    for shift_type, label in [(ShiftType.DAY, "주간합"), (ShiftType.NIGHT, "야간합")]:
        r = len(result.employees) + 2
        ws.cell(r, 1, label).font = Font(bold=True)
        for col, d in enumerate(days, 2):
            cnt = sum(1 for e in result.employees
                      if result.get_shift(e.name, d) == shift_type)
            ws.cell(r, col, cnt).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 10
    for col in range(2, len(days) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 4

    wb.save(filepath)
    return filepath
